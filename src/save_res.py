"""

"""
import os
from openpyxl import Workbook, load_workbook

def save_experiment_data(file_name, data):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "result"  #type:ignore 
        ws.append(["index"])  #type:ignore 
        indicators = ["profit", "waiting time", "detour time"]
        for idx, indicator in enumerate(indicators, start=2):  
            ws.cell(row=idx, column=1, value=indicator) #type:ignore 
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    max_column = ws.max_column + 1  #type:ignore 

    ws.cell(row=1, column=max_column, value=max_column - 1)  #type:ignore 

    for idx, value in enumerate(data, start=2):  
        ws.cell(row=idx, column=max_column, value=value) #type:ignore 

    wb.save(file_name)
    print(f"result save to {file_name}")


